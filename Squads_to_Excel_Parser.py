from openpyxl import Workbook
from api_swgoh_help import api_swgoh_help, Settings
from openpyxl.styles import Alignment, Border, Side

# api_swgoh_help by MarTrepodi https://github.com/MarTrepodi/api_swgoh_help #

# allycode = int(input('Введите код союзника: '))  # Your AllyCode

# in order to find units names you have to use get_your_roster func and go through dict

units_list_GRIEVOUS = {'B1BATTLEDROIDV2': 'B1', 'DROIDEKA': 'Дроидека', 'GRIEVOUS': 'Гривус',
                       'MAGNAGUARD': 'Магна', 'B2SUPERBATTLEDROID': 'B2'}
units_list_GEO = {'POGGLETHELESSER': 'Поггль', 'GEONOSIANBROODALPHA': 'Альфа', 'GEONOSIANSPY': 'Шпион',
                  'GEONOSIANSOLDIER': 'Солдат', 'SUNFAC': 'Сан Фак'}

units_list = {'HANSOLO': 'Хан'}


# Client = Settings(input('Введите логин от api.swgoh.help: '), input('Введите пароль: '))


# Sign Up on api.swgoh.help/ and use login, pass as parameters -> check MarTrepodi`s api_swgoh_help module

class SwgohData:
    def __init__(self):
        self.allycode = int(input('Введите ваш код союзника: '))
        self.client = Settings(input('Введите логин от api.swgoh.help: '), input('Введите пароль: '))


class GuildSquadsData(SwgohData):

    def __init__(self, units_list: dict):
        print(f'Инициализирую создание таблицы по переданному списку юнитов')
        super().__init__()
        self.wb = Workbook()
        self.sheet = self.wb.active
        self.guild_allycodes = self.get_guild_allycodes()
        self.units_list = units_list
        self.guild_units = self.fetch_guild_units()

    def get_guild_allycodes(self):
        try:
            print('Запрашиваю коды союзников членов гильдии')
            guild_allycodes = {}
            data = api_swgoh_help(self.client).fetchGuilds(self.allycode)
            for i in data[0]['roster']:  # Get AllyCodes from your team
                guild_allycodes[i['allyCode']] = i['name']
            return guild_allycodes
        except:
            raise Exception('Возникла ошибка, проверьте точность введенных вами данных')

    def fetch_guild_units(self):
        print('Собираю список юнитов членов гильдии')
        # put get_guild_allycodes function as parameter with your allycode or get allycodes first
        data = api_swgoh_help(self.client).fetchRoster(list(self.guild_allycodes.keys()))
        res_dict = {}
        for i in self.guild_allycodes.keys():
            res_dict[i] = []
        for i in data:
            for unit in i:
                if unit in self.units_list:
                    res_dict[i[unit][0]['allyCode']].append({unit: i[unit]})
        nicknamed_dict = {self.guild_allycodes[i]: res_dict[i] for i in res_dict}
        # changing allycodes to nicknames for Excel representation
        return nicknamed_dict

    def find_column(self, search_value):
        for i in range(len(self.units_list) + 2):
            if self.sheet.cell(1, 2 + i).value == search_value:
                return 2 + i

    @staticmethod
    def get_needed_data(hero_data):
        needed_data = ['gp', 'starLevel', 'gearLevel', 'zetas']
        data = [hero_data[i] for i in needed_data if i in hero_data]
        if len(hero_data['zetas']) > 0:
            return f'ГМ: {data[0]}, звезд: {data[1]}, Гир: {data[2]}, Кол-во дзет: {len(data[3])}'
        return f'ГМ: {data[0]}, звезд: {data[1]}, Гир: {data[2]}'

    @staticmethod
    def make_styles(sheet):
        dims = {}
        for row in sheet.rows:
            for cell in row:
                if cell.value is not None:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
                cell.alignment = Alignment(horizontal='center', vertical='center')
                thin = Side(border_style='thin')
                cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        for col, value in dims.items():
            sheet.column_dimensions[col].width = value

    def write_to_sheet(self, needed_nicknames=None, seven=None):  # lend units_list you want to get information about
        print('Начинаю запись в файл')
        self.sheet.cell(1, 1).value = 'Name'
        for i in range(len(self.units_list)):
            self.sheet.cell(1, i + 2).value = list(self.units_list.keys())[i]
        self.sheet.cell(1, len(self.units_list) + 2).value = 'Comments'
        self.sheet.cell(1, len(self.units_list) + 3).value = 'Итог'
        self.sheet.cell(1, len(self.units_list) + 4).value = 'Проверка'  # базовая разметка
        _working_dict = {nickname: self.guild_units[nickname] for nickname in self.guild_units if
                         (needed_nicknames is None or nickname in needed_nicknames)}
        # print(_working_dict)
        for indx, nickname in enumerate(_working_dict):
            # indx used to choose right row, nickname for further searching and writting nickname in Name column
            self.sheet.cell(2 + indx, 1).value = nickname
            for hero in _working_dict[nickname]:  # getting heroes list for every ally
                hero_name = list(hero.keys())[0]
                if hero_name in self.units_list:  # search if your ally HAS particular unit
                    if seven is None:
                        column = self.find_column(hero_name)
                        self.sheet.cell(2 + indx, column).value = self.get_needed_data(
                            _working_dict[nickname][_working_dict[nickname].index(hero)][hero_name][0])
                    elif seven is not None and hero[hero_name][0]['starLevel'] != 7:
                        column = self.find_column(hero_name)
                        self.sheet.cell(2 + indx, column).value = self.get_needed_data(
                            _working_dict[nickname][_working_dict[nickname].index(hero)][hero_name][0])
        for i in range(len(self.units_list)):
            self.sheet.cell(1, i + 2).value = self.units_list[list(self.units_list.keys())[i]]
        self.make_styles(self.sheet)
        # for i in guild_units:
        #     sheet.cell()
        self.wb.save(f'/Users/{input("Введите имя пользователя в Windows: ")}/Desktop/SWGOH_SQUAD_KENOBI.xlsx')
        print('Готово! Проверьте ваш рабочий стол')

    def write_data_by_nicknames(self):
        needed_nicknames = input('Введите через пробел никнеймы, по которым необходимо получить информацию: ').split()
        return self.write_to_sheet(needed_nicknames)


class SelfData(SwgohData):
    def __init__(self):
        super().__init__()
        self.units_data = api_swgoh_help(self.client).fetchUnits([self.allycode])

    def get_data(self):
        data = api_swgoh_help(self.client).fetchGuilds([self.allycode])
        return data

    def get_mods_on_hero(self):
        mod_atts = {'set': {1: 'ХП', 2: 'Оборона', 3: 'Крит. урон', 4: 'Шанс крит. урона', 5: 'Стойкость', 6: 'Атака',
                            7: 'Эффективность', 8: 'Скорость'}, 'tier': {1: 'Обычный', 2: 'Необычный', 3: 'Редкий '}}
        hero = input('Введите название юнита, по которому необходимо получить информацию: ')
        if hero in self.units_data:
            for i in self.units_data[hero][0]['mods']:
                print(i)


a = GuildSquadsData({'GENERALKENOBI': 'Кеноби'})
a.write_to_sheet(seven=True)
# a = SelfData()
# a.get_mods_on_hero()
# write_to_sheet(units_list_GEO)
# print(a)
# fetch_guild_units(a)
# a = GuildSquadsData(units_list)
# a.write_to_sheet()
# a.write_data_by_nicknames()
